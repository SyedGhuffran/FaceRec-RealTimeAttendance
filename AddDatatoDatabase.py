import firebase_admin
from firebase_admin import credentials
from firebase_admin import db
from openpyxl import load_workbook

cred = credentials.Certificate("serviceAccountKey.json")
firebase_admin.initialize_app(cred,{
    'databaseURL' : "Add Your Database Cred here"
})
ref = db.reference('Employees')

# Load data from Excel file
wb = load_workbook(filename='employees.xlsx')
ws = wb.active

# Convert data to a dictionary
data = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    key = str(row[0])
    value = {
        'name': row[1],
        'designation': row[2],
        'department': row[3],
        'work_location': row[4],
        'cnic': row[5],
        'mobile': row[6]
    }
    data[key] = value

# Add data to Firebase Realtime Database
for key, value in data.items():
    ref.child(key).set(value)
